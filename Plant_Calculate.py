import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
from pandastable import Table
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import sqlite3

class DataSayuranApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Data Sayuran")
        self.configure(bg="#FDFDFD")

        self.db_conn = None
        self.file_path = ""
        self.table_frame = None
        self.table = None

        self.create_widgets()

    def create_widgets(self):
        # Create a top label
        label = tk.Label(self, text="Data Sayuran", font=("Arial", 24), bg="#FDFDFD")
        label.grid(row=0, column=0, columnspan=4, padx=10, pady=20)

        # Create a sub-label
        sub_label = tk.Label(self, text="Analisis Produksi Sayuran \n (Buncis)", font=("Arial", 16), bg="#FDFDFD")
        sub_label.grid(row=1, column=0, columnspan=4, padx=10, pady=10)

        # Create buttons with custom styling
        button_style = {"font": ("Arial", 14), "width": 15, "height": 2}
        browse_button = tk.Button(self, text="Pilih File Excel", command=self.browse_file, **button_style)
        browse_button.grid(row=2, column=0, padx=10, pady=10)

        add_data_button = tk.Button(self, text="Tambah Data", command=self.add_data, **button_style)
        add_data_button.grid(row=2, column=1, padx=10, pady=10)

        statistics_button = tk.Button(self, text="Statistik", command=self.show_statistics, **button_style)
        statistics_button.grid(row=2, column=2, padx=10, pady=10)

        comparison_button = tk.Button(self, text="Grafik Perbandingan", command=self.show_comparison_chart, **button_style)
        comparison_button.grid(row=2, column=3, padx=10, pady=10)

        # Menu bar
        menubar = tk.Menu(self)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Keluar", command=self.exit_application)
        menubar.add_cascade(label="File", menu=file_menu)

        # Add "View Database" menu item
        view_menu = tk.Menu(menubar, tearoff=0)
        view_menu.add_command(label="View Database", command=self.view_database)
        menubar.add_cascade(label="View", menu=view_menu)

        self.config(menu=menubar)

    def browse_file(self):
        self.file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if self.file_path:
            self.load_data()

    def load_data(self):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active

        # Mengambil jumlah kolom dan baris dari file Excel
        num_rows = sheet.max_row
        num_cols = sheet.max_column

        # Membaca data dari setiap sel, mulai dari baris kedua
        data = []
        for row in sheet.iter_rows(min_row=2, max_row=num_rows, min_col=1, max_col=num_cols, values_only=True):
            data.append(row)

        # Clear the existing table frame if it exists
        if self.table_frame:
            self.table_frame.destroy()

        # Membuat tabel dari data Excel menggunakan PandasTable
        self.table_frame = tk.Frame(self)
        self.table_frame.grid(row=3, column=0, columnspan=4, padx=10, pady=10)

        self.table = Table(self.table_frame, dataframe=pd.DataFrame(data))
        self.table.show()

        self.save_data_to_database(data)

    def save_data_to_database(self, rows):
        if self.db_conn:
            self.db_conn.close()

        self.db_conn = sqlite3.connect(":memory:")  # Menggunakan database SQLite di dalam memori
        cursor = self.db_conn.cursor()

        # Membuat tabel "sayuran" dengan kolom-kolom numerik
        column_names = [f"col{i}" for i in range(1, len(rows[0]) + 1)]
        create_table_query = f"CREATE TABLE sayuran ({','.join(column_names)});"
        cursor.execute(create_table_query)

        # Memasukkan baris-baris data ke dalam tabel
        insert_data_query = f"INSERT INTO sayuran VALUES ({','.join(['?'] * len(column_names))});"
        cursor.executemany(insert_data_query, rows)

        self.db_conn.commit()

    def view_database(self):
        if self.db_conn:
            cursor = self.db_conn.cursor()
            cursor.execute("SELECT * FROM sayuran")
            rows = cursor.fetchall()

            # Create a new window for displaying the database view
            db_view_window = tk.Toplevel(self)
            db_view_window.title("Database View")

            # Create a frame to hold the table view
            table_frame = tk.Frame(db_view_window)
            table_frame.pack(padx=10, pady=10)

            # Create a PandasTable instance to display the database content
            table = Table(table_frame, dataframe=pd.DataFrame(rows))
            table.show()
        else:
            messagebox.showerror("Error", "Database connection is not established.")

    def add_data(self):
        if self.file_path:
            new_file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
            if new_file_path:
                wb = openpyxl.load_workbook(new_file_path)
                sheet = wb.active
                new_data = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    new_data.append(row)
                df = pd.DataFrame(new_data, columns=sheet[1])

                # Tambahkan data ke database SQLite
                cursor = self.db_conn.cursor()
                insert_data_query = f"INSERT INTO sayuran VALUES ({','.join(['?']*len(df.columns))});"
                cursor.executemany(insert_data_query, df.values)

                self.db_conn.commit()
                self.load_data()
        else:
            messagebox.showerror("Error", "Pilih file Excel terlebih dahulu.")

    def show_statistics(self):
        if self.file_path:
            df = pd.read_excel(self.file_path)
            buncis_data = df['Buncis (Ton)']
            max_production = buncis_data.max()
            min_production = buncis_data.min()
            average_production = buncis_data.mean()
            messagebox.showinfo("Statistik Produksi Buncis", f"Produksi Maksimal: {max_production}\nProduksi Minimal: {min_production}\nRata-rata Produksi: {average_production}")
        else:
            messagebox.showerror("Error", "Pilih file Excel terlebih dahulu.")

    def show_comparison_chart(self):
        if self.file_path:
            df = pd.read_excel(self.file_path)
            provinces = df['Provinsi'].astype(str)  # Convert values to strings
            buncis_data = df['Buncis (Ton)']

            # Create a new window for displaying the chart
            chart_window = tk.Toplevel(self)
            chart_window.title("Grafik Perbandingan")

            # Create a figure and axis for the comparison chart
            fig, ax = plt.subplots()
            ax.bar(provinces, buncis_data)
            ax.set_xlabel('Provinsi')
            ax.set_ylabel('Produksi Buncis (Ton)')
            ax.set_title('Perbandingan Produksi Buncis Tiap Provinsi')
            ax.tick_params(axis='x', rotation=90)

            # Embed the chart in the Tkinter window
            canvas = FigureCanvasTkAgg(fig, master=chart_window)
            canvas.draw()
            canvas.get_tk_widget().pack()
        else:
            messagebox.showerror("Error", "Pilih file Excel terlebih dahulu.")

    def exit_application(self):
        self.destroy()

if __name__ == '__main__':
    app = DataSayuranApp()
    app.mainloop()
